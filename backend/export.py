import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")


class ReportExporter:

    DARK_BLUE = "1a1f3a"
    ACCENT_BLUE = "00d4ff"
    GOLD = "ffd700"
    WHITE = "ffffff"
    LIGHT_GRAY = "f0f4f8"
    MID_GRAY = "8892b0"

    def _header_fill(self):
        return PatternFill("solid", fgColor=self.DARK_BLUE)

    def _accent_fill(self):
        return PatternFill("solid", fgColor="1e3a5f")

    def _header_font(self, size=11):
        return Font(bold=True, color=self.WHITE, size=size, name="Calibri")

    def _body_font(self, size=10):
        return Font(size=size, name="Calibri")

    def _thin_border(self):
        s = Side(style="thin", color="d0d7de")
        return Border(left=s, right=s, top=s, bottom=s)

    def _style_header_row(self, ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self._header_fill()
            cell.font = self._header_font()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._thin_border()

    def _style_data_row(self, ws, row, cols, alternate=False):
        fill = PatternFill("solid", fgColor=self.LIGHT_GRAY if alternate else self.WHITE)
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = self._body_font()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._thin_border()

    def _write_title(self, ws, title: str, row: int, cols: int, size=16):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(cols, 10))
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(bold=True, size=size, color=self.DARK_BLUE, name="Calibri")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def to_excel(self, result: dict, output_path_or_source: str, output_path: str = None):
        # Handle both calling conventions: to_excel(result, out) and to_excel(result, src, out)
        if output_path is None:
            output_path = output_path_or_source
        wb = Workbook()
        wb.remove(wb.active)

        self._sheet_cover(wb, result)
        self._sheet_summary(wb, result)
        self._sheet_statistics(wb, result)
        self._sheet_correlations(wb, result)
        self._sheet_outliers(wb, result)
        self._sheet_categorical(wb, result)
        self._sheet_insights(wb, result)
        src_path = output_path_or_source if output_path is not None else ""
        self._sheet_raw_data(wb, src_path, result["metadata"])
        self._sheet_charts(wb, result)

        wb.save(output_path)

    def to_powerbi_excel(self, result: dict, source_path: str, output_path: str):
        """Power BI optimized Excel: structured tables, proper headers, no merged cells."""
        wb = Workbook()
        wb.remove(wb.active)

        # Raw data sheet (most important for PBI)
        self._sheet_raw_data(wb, source_path, result["metadata"], sheet_name="Data")
        # Stats table
        self._sheet_statistics(wb, result, sheet_name="Statistics")
        # Correlation table
        self._sheet_correlations(wb, result, sheet_name="Correlations")
        # Categorical
        self._sheet_categorical(wb, result, sheet_name="Categories")
        # Insights as table
        self._sheet_insights(wb, result, sheet_name="Insights")
        # PBI Guide
        self._sheet_pbi_guide(wb)

        wb.save(output_path)

    def _sheet_cover(self, wb, result):
        ws = wb.create_sheet("📊 Report Cover")
        meta = result["metadata"]

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 30

        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 60
        ws.row_dimensions[3].height = 20

        ws.merge_cells("B2:C2")
        cell = ws["B2"]
        cell.value = "📊 DataVision Pro — Analytics Report"
        cell.font = Font(bold=True, size=22, color=self.DARK_BLUE, name="Calibri")
        cell.alignment = Alignment(horizontal="left", vertical="center")

        info = [
            ("File", meta.get('filename', meta.get('label','Dataset'))),
            ("Analyzed", meta["analyzed_at"][:19].replace("T", " ")),
            ("Rows", f"{meta['rows']:,}"),
            ("Columns", meta["columns"]),
            ("Numeric Columns", len(meta.get('numeric_columns',meta.get('numeric_cols',[])))),
            ("Categorical Columns", len(meta.get('categorical_columns',meta.get('categorical_cols',[])))),
            ("Quality", f"{result['profile']['completeness_pct']}% complete"),
        ]
        for i, (label, val) in enumerate(info):
            row = 4 + i
            ws.cell(row=row, column=2).value = label
            ws.cell(row=row, column=2).font = Font(bold=True, color=self.MID_GRAY, name="Calibri")
            ws.cell(row=row, column=3).value = val
            ws.cell(row=row, column=3).font = Font(size=11, name="Calibri")

    def _sheet_summary(self, wb, result):
        ws = wb.create_sheet("📋 Summary")
        profile = result["profile"]

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

        self._write_title(ws, "DATA PROFILE SUMMARY", 1, 2, 14)

        headers = [("Metric", "Value")]
        data = [
            ("Total Rows", f"{_rows:,}"),
            ("Total Columns", _cols),
            ("Duplicate Rows", f"{_dups:,}"),
            ("Data Completeness", f"{_comp}%"),
            ("Memory Usage (MB)", round(float(_mem), 3)),
        ]
        if _miss:
            for col, cnt in list(_miss.items())[:5]:
                data.append((f"Missing: {col}", f"{cnt:,} ({_misspct.get(col, 0):.1f}%)"))

        for r, (h1, h2) in enumerate(headers, 3):
            ws.cell(row=r, column=1).value = h1
            ws.cell(row=r, column=2).value = h2
            self._style_header_row(ws, r, 2)

        for r, (label, val) in enumerate(data, 4):
            ws.cell(row=r, column=1).value = label
            ws.cell(row=r, column=2).value = val
            self._style_data_row(ws, r, 2, r % 2 == 0)

        # KPIs
        ws.cell(row=4 + len(data) + 2, column=1).value = "KEY METRICS"
        ws.cell(row=4 + len(data) + 2, column=1).font = Font(bold=True, size=12, color=self.DARK_BLUE)

        for i, kpi in enumerate(result.get("kpis", [])):
            row = 4 + len(data) + 3 + i
            ws.cell(row=row, column=1).value = kpi["label"]
            ws.cell(row=row, column=2).value = kpi["value"]
            self._style_data_row(ws, row, 2, i % 2 == 0)

    def _sheet_statistics(self, wb, result, sheet_name="📈 Statistics"):
        ws = wb.create_sheet(sheet_name)
        stats = result.get("statistics", {})
        if not stats:
            ws["A1"] = "No numeric columns found."
            return

        self._write_title(ws, "DESCRIPTIVE STATISTICS", 1, 14, 14)

        headers = ["Column", "Count", "Mean", "Median", "Std Dev", "Variance",
                   "Min", "Max", "Range", "Q1", "Q3", "IQR", "Skewness", "Kurtosis", "CV%"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c).value = h
        self._style_header_row(ws, 3, len(headers))

        for r, (col, s) in enumerate(stats.items(), 4):
            row_data = [col, s["count"], s["mean"], s["median"], s["std"], s["variance"],
                        s["min"], s["max"], s["range"], s["q1"], s["q3"], s["iqr"],
                        s["skewness"], s["kurtosis"], s.get("cv")]
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c).value = val
            self._style_data_row(ws, r, len(headers), r % 2 == 0)

        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14
        ws.column_dimensions["A"].width = 25

    def _sheet_correlations(self, wb, result, sheet_name="🔗 Correlations"):
        ws = wb.create_sheet(sheet_name)
        corr = result.get("correlations", {})
        matrix = corr.get("matrix", {})
        strong = corr.get("strong_pairs", [])

        self._write_title(ws, "CORRELATION ANALYSIS", 1, 10, 14)

        if matrix:
            cols = list(matrix.keys())
            ws.cell(row=3, column=1).value = "Column"
            ws.cell(row=3, column=1).font = self._header_font()
            ws.cell(row=3, column=1).fill = self._header_fill()
            for c, col in enumerate(cols, 2):
                ws.cell(row=3, column=c).value = col
                ws.cell(row=3, column=c).font = self._header_font(9)
                ws.cell(row=3, column=c).fill = self._header_fill()
                ws.cell(row=4 + c - 2, column=1).value = col
                ws.cell(row=4 + c - 2, column=1).font = Font(bold=True, size=10, name="Calibri")
            for r, (row_col, vals) in enumerate(matrix.items(), 4):
                for c, col in enumerate(cols, 2):
                    val = vals.get(col, "")
                    cell = ws.cell(row=r, column=c)
                    cell.value = val if val != "" else ""
                    if isinstance(val, (int, float)):
                        cell.number_format = "0.000"
                        abs_val = abs(val)
                        if row_col == col:
                            cell.fill = PatternFill("solid", fgColor="c8e6c9")
                        elif abs_val > 0.8:
                            cell.fill = PatternFill("solid", fgColor="ef9a9a")
                        elif abs_val > 0.5:
                            cell.fill = PatternFill("solid", fgColor="ffe082")
                        else:
                            cell.fill = PatternFill("solid", fgColor=self.WHITE)
                    cell.border = self._thin_border()
                    cell.alignment = Alignment(horizontal="center")

        if strong:
            start_row = 4 + len(matrix) + 3
            ws.cell(row=start_row, column=1).value = "STRONG CORRELATIONS"
            ws.cell(row=start_row, column=1).font = Font(bold=True, size=12, color=self.DARK_BLUE)
            headers = ["Column 1", "Column 2", "Correlation", "Strength", "Direction", "Interpretation"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=start_row + 1, column=c).value = h
            self._style_header_row(ws, start_row + 1, len(headers))
            for r, pair in enumerate(strong, start_row + 2):
                row_data = [pair.get("col1",""), pair.get("col2",""), pair.get("r", pair.get("correlation", 0)),
                            pair.get("strength",""), pair.get("direction",""), pair.get("interpretation","")]
                for c, val in enumerate(row_data, 1):
                    ws.cell(row=r, column=c).value = val
                self._style_data_row(ws, r, len(headers), r % 2 == 0)

        for c in range(1, 20):
            ws.column_dimensions[get_column_letter(c)].width = 16

    def _sheet_outliers(self, wb, result):
        ws = wb.create_sheet("⚠️ Outliers")
        outliers = result.get("outliers", {})

        self._write_title(ws, "OUTLIER DETECTION (IQR Method)", 1, 7, 14)
        headers = ["Column", "Outlier Count", "Outlier %", "Lower Bound", "Upper Bound", "Severity", "Sample Values"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c).value = h
        self._style_header_row(ws, 3, len(headers))

        for r, (col, data) in enumerate(outliers.items(), 4):
            lb  = data.get("lower_bound", data.get("lb", ""))
            ub  = data.get("upper_bound", data.get("ub", ""))
            cnt = data.get("count", 0)
            pct = data.get("pct", data.get("outlier_pct", 0))
            sev = data.get("severity", "Low")
            ext = data.get("extreme_values", data.get("extremes", []))
            row_data = [col, cnt, f"{pct}%", lb, ub, sev, str(ext[:3])]
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c).value = val
            self._style_data_row(ws, r, len(headers), r % 2 == 0)
            sev_cell = ws.cell(row=r, column=6)
            if sev == "High":
                sev_cell.fill = PatternFill("solid", fgColor="ef9a9a")
                sev_cell.font = Font(bold=True, color="c62828")
            elif sev == "Medium":
                sev_cell.fill = PatternFill("solid", fgColor="ffe082")

        widths = [25, 15, 12, 15, 15, 12, 40]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

    def _sheet_categorical(self, wb, result, sheet_name="🏷️ Categories"):
        ws = wb.create_sheet(sheet_name)
        cat = result.get("categorical_analysis", {})

        self._write_title(ws, "CATEGORICAL ANALYSIS", 1, 5, 14)
        current_row = 3
        for col, data in cat.items():
            ws.cell(row=current_row, column=1).value = f"Column: {col}"
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color=self.DARK_BLUE)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            current_row += 1
            info = [("Unique Values", data.get("unique_count", data.get("unique",0))),
                    ("Mode", data.get("mode","")),
                    ("Cardinality", data.get("cardinality","")),
                    ("Missing", data.get("missing_count", data.get("missing",0)))]
            for label, val in info:
                ws.cell(row=current_row, column=1).value = label
                ws.cell(row=current_row, column=2).value = str(val)
                current_row += 1
            headers = ["Value", "Count", "Percentage"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=current_row, column=c).value = h
            self._style_header_row(ws, current_row, 3)
            current_row += 1
            for r, item in enumerate(data["top_values"]):
                ws.cell(row=current_row, column=1).value = item["value"]
                ws.cell(row=current_row, column=2).value = item["count"]
                ws.cell(row=current_row, column=3).value = f"{item['pct']}%"
                self._style_data_row(ws, current_row, 3, r % 2 == 0)
                current_row += 1
            current_row += 2

        for c in range(1, 6):
            ws.column_dimensions[get_column_letter(c)].width = 25

    def _sheet_insights(self, wb, result, sheet_name="💡 Insights"):
        ws = wb.create_sheet(sheet_name)
        insights = result.get("insights", [])

        self._write_title(ws, "AI-GENERATED DATA INSIGHTS", 1, 5, 14)
        headers = ["#", "Type", "Priority", "Title", "Description"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c).value = h
        self._style_header_row(ws, 3, len(headers))

        priority_colors = {"high": "ef9a9a", "medium": "ffe082", "low": "c8e6c9"}
        for r, ins in enumerate(insights, 4):
            row_data = [r - 3, ins["type"].upper(), ins["priority"].upper(), ins["title"], ins.get('insight', ins.get('description', ''))]
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c)
                cell.value = val
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            self._style_data_row(ws, r, len(headers), r % 2 == 0)
            prio_cell = ws.cell(row=r, column=3)
            prio_color = priority_colors.get(ins["priority"], "ffffff")
            prio_cell.fill = PatternFill("solid", fgColor=prio_color)
            prio_cell.font = Font(bold=True, size=10)

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 70
        ws.row_dimensions[1].height = 30
        for r in range(4, 4 + len(insights)):
            ws.row_dimensions[r].height = 45

    def _sheet_raw_data(self, wb, source_path: str, meta: dict, sheet_name="📂 Raw Data"):
        ws = wb.create_sheet(sheet_name)
        try:
            fname = meta.get('filename', meta.get('label','Dataset'))
            ext = fname.split(".")[-1].lower() if '.' in fname else 'csv' 
            if ext == "csv":
                df = pd.read_csv(source_path, nrows=10000)
            elif ext in ["xlsx", "xls"]:
                df = pd.read_excel(source_path, nrows=10000)
            elif ext == "json":
                df = pd.read_json(source_path)
                df = df.head(10000)
            else:
                ws["A1"] = "Source file not available for raw data export."
                return
        except Exception as e:
            ws["A1"] = f"Could not load raw data: {str(e)}"
            return

        for c, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=c)
            cell.value = str(col)
        self._style_header_row(ws, 1, len(df.columns))

        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, val in enumerate(row, 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    if pd.isna(val):
                        cell.value = None
                    elif isinstance(val, (np.integer,)):
                        cell.value = int(val)
                    elif isinstance(val, (np.floating,)):
                        cell.value = float(val)
                    else:
                        cell.value = str(val)
                except Exception:
                    ws.cell(row=r, column=c).value = str(val)
            if r % 2 == 0:
                for c in range(1, len(df.columns) + 1):
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=self.LIGHT_GRAY)

        for c in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18

    def _sheet_charts(self, wb, result):
        ws = wb.create_sheet("📉 Chart Data")
        chart_data = result.get("chart_data", {})

        self._write_title(ws, "CHART DATA (Use this sheet for Power BI visuals)", 1, 8, 13)
        current_row = 3

        for chart_key, chart in chart_data.items():
            ws.cell(row=current_row, column=1).value = chart.get("title", chart_key)
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color=self.DARK_BLUE)
            current_row += 1

            data = chart.get("data", [])
            if not data:
                current_row += 2
                continue

            keys = list(data[0].keys())
            for c, k in enumerate(keys, 1):
                ws.cell(row=current_row, column=c).value = k
            self._style_header_row(ws, current_row, len(keys))
            current_row += 1

            for item in data:
                for c, k in enumerate(keys, 1):
                    ws.cell(row=current_row, column=c).value = item.get(k)
                current_row += 1
            current_row += 2

    def _sheet_pbi_guide(self, wb):
        ws = wb.create_sheet("🔗 Open in Power BI")
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 80

        ws.merge_cells("B2:B20")
        guide = """HOW TO OPEN THIS REPORT IN POWER BI DESKTOP

Step 1: Open Power BI Desktop (free download from microsoft.com/power-bi)

Step 2: Click "Get Data" → "Excel Workbook"

Step 3: Select this file (DataVision_PowerBI_Ready.xlsx)

Step 4: In the Navigator, select these sheets:
  ✓ Data          → Your raw dataset
  ✓ Statistics    → Descriptive statistics table
  ✓ Correlations  → Correlation matrix
  ✓ Categories    → Categorical breakdowns
  ✓ Insights      → Generated insights table

Step 5: Click "Transform Data" to open Power Query Editor
  • Set correct data types for each column
  • Filter out header rows if needed
  • Click "Close & Apply"

Step 6: Create visualizations:
  • Bar chart: Categories sheet → Value + Count fields
  • Line chart: Data sheet → Date + numeric columns
  • KPI cards: Statistics sheet → Mean, Sum columns
  • Matrix: Correlations sheet → heatmap visualization

Step 7: Save as .pbix file (File → Save As → Power BI Report)

That's it! Your real DataVision Pro analysis is now in Power BI Desktop.
"""
        ws["B2"].value = guide
        ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["B2"].font = Font(size=12, name="Calibri")
        ws.row_dimensions[2].height = 400

    def to_pdf(self, result: dict, output_path: str):
        doc = SimpleDocTemplate(
            output_path, pagesize=A4,
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=2*cm, bottomMargin=2*cm
        )
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            "Title", parent=styles["Heading1"],
            fontSize=24, textColor=colors.HexColor("#1a1f3a"),
            spaceAfter=6, alignment=TA_LEFT
        )
        subtitle_style = ParagraphStyle(
            "Subtitle", parent=styles["Normal"],
            fontSize=11, textColor=colors.HexColor("#8892b0"),
            spaceAfter=20
        )
        h2_style = ParagraphStyle(
            "H2", parent=styles["Heading2"],
            fontSize=14, textColor=colors.HexColor("#1a1f3a"),
            spaceBefore=16, spaceAfter=8
        )
        body_style = ParagraphStyle(
            "Body", parent=styles["Normal"],
            fontSize=10, leading=14,
            textColor=colors.HexColor("#333333")
        )

        meta = result["metadata"]
        story.append(Paragraph("📊 DataVision Pro", title_style))
        story.append(Paragraph(f"Analytics Report — {meta.get('filename', meta.get('label','Dataset'))}", subtitle_style))
        story.append(Paragraph(f"Generated: {meta['analyzed_at'][:19].replace('T', ' ')}", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#00d4ff")))
        story.append(Spacer(1, 0.3*inch))

        # KPIs
        story.append(Paragraph("KEY METRICS", h2_style))
        kpis = result.get("kpis", [])
        if kpis:
            kpi_data = [["Metric", "Value"]] + [[k["label"], k["value"]] for k in kpis]
            kpi_table = Table(kpi_data, colWidths=[3*inch, 2.5*inch])
            kpi_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1f3a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f0f4f8"), colors.white]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7de")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWHEIGHT", (0, 0), (-1, -1), 22),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(kpi_table)
        story.append(Spacer(1, 0.2*inch))

        # Statistics
        stats = result.get("statistics", {})
        if stats:
            story.append(Paragraph("DESCRIPTIVE STATISTICS", h2_style))
            stat_rows = [["Column", "Mean", "Median", "Std Dev", "Min", "Max", "Skewness"]]
            for col, s in list(stats.items())[:10]:
                stat_rows.append([col, f"{s['mean']:.3f}", f"{s['median']:.3f}",
                                  f"{s['std']:.3f}", f"{s['min']:.3f}", f"{s['max']:.3f}",
                                  f"{s['skewness']:.3f}"])
            stat_table = Table(stat_rows, colWidths=[1.8*inch] + [0.95*inch]*6)
            stat_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1f3a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f0f4f8"), colors.white]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7de")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(stat_table)
            story.append(Spacer(1, 0.2*inch))

        # Insights
        insights = result.get("insights", [])
        if insights:
            story.append(PageBreak())
            story.append(Paragraph("💡 DATA INSIGHTS", h2_style))
            for ins in insights:
                priority_colors_map = {"high": "#c62828", "medium": "#e65100", "low": "#1b5e20"}
                color = priority_colors_map.get(ins["priority"], "#333")
                story.append(Paragraph(
                    f'<font color="{color}"><b>[{ins["priority"].upper()}] {ins["title"]}</b></font>',
                    ParagraphStyle("InsTitle", parent=body_style, spaceAfter=2, spaceBefore=10)
                ))
                story.append(Paragraph(ins.get('insight', ins.get('description', '')), body_style))
            story.append(Spacer(1, 0.2*inch))

        # Correlations
        corr = result.get("correlations", {})
        strong_pairs = corr.get("strong_pairs", [])
        if strong_pairs:
            story.append(Paragraph("🔗 STRONG CORRELATIONS", h2_style))
            corr_rows = [["Column 1", "Column 2", "Correlation", "Strength", "Direction"]]
            for pair in strong_pairs[:10]:
                corr_rows.append([pair.get("col1",""), pair.get("col2",""),
                                  f"{pair.get('r', pair.get('correlation', pair.get('r_squared',0))):.4f}", pair["strength"], pair["direction"]])
            corr_table = Table(corr_rows, colWidths=[1.8*inch, 1.8*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            corr_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1f3a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#f0f4f8"), colors.white]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d0d7de")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(corr_table)

        doc.build(story)
